import streamlit as st
import pandas as pd
import openpyxl
import io

# ==========================================
# KONFIGURASI & TAMPILAN MINIMALIS ADAPTIF
# ==========================================
st.set_page_config(page_title="PLN Wellness Sync", layout="centered", page_icon="⚡")

st.markdown("""
    <style>
        /* Header yang tajam dan elegan */
        .header-container {
            border-bottom: 1px solid rgba(150, 150, 150, 0.2);
            padding-top: 2rem;
            padding-bottom: 1.5rem;
            margin-bottom: 2.5rem;
        }
        .header-title {
            font-size: 2.4rem;
            font-weight: 300;
            letter-spacing: -0.5px;
            margin: 0;
        }
        .header-subtitle {
            font-size: 0.95rem;
            opacity: 0.7;
            margin-top: 0.5rem;
            font-weight: 400;
        }
        
        /* Styling teks sub-bab */
        h3 {
            font-size: 1.1rem !important;
            font-weight: 600 !important;
            letter-spacing: 1px !important;
            text-transform: uppercase;
            margin-bottom: 1rem !important;
        }

        /* Tombol outline artistik */
        .stButton>button {
            background-color: transparent !important;
            border: 1px solid #00A2E9 !important;
            border-radius: 4px !important;
            padding: 0.6rem 2rem !important;
            font-weight: 500 !important;
            letter-spacing: 1px !important;
            text-transform: uppercase !important;
            font-size: 0.85rem !important;
            transition: all 0.3s ease !important;
            width: 100%;
            margin-top: 2rem;
        }
        .stButton>button:hover {
            background-color: #00A2E9 !important;
            color: white !important;
        }
        
        /* Footer ala signature */
        .footer-text {
            margin-top: 5rem;
            padding-top: 1rem;
            border-top: 1px solid rgba(150, 150, 150, 0.2);
            text-align: right;
            font-size: 0.75rem;
            opacity: 0.6;
            font-family: 'Courier New', Courier, monospace;
            letter-spacing: 0.5px;
        }
        .footer-text span {
            font-weight: bold;
        }
    </style>

    <div class="header-container">
        <h1 class="header-title">PLN Wellness Data Sync.</h1>
        <div class="header-subtitle">Rekapitulasi dan pembaruan klasemen mingguan otomatis.</div>
    </div>
""", unsafe_allow_html=True)


# ==========================================
# FORM UPLOAD DATA
# ==========================================
st.markdown("### 01. MASTER DATA")
master_file = st.file_uploader("Upload Excel Master (Original)", type=["xlsx"])

st.write("<br>", unsafe_allow_html=True) # Jarak aman biar markdown nggak rusak
st.markdown("### 02. UPDATE MINGGUAN")
col1, col2 = st.columns(2)
with col1:
    file_active = st.file_uploader("active inactive.xlsx", type=["xlsx"])
    file_kcal = st.file_uploader("Kcal.xlsx", type=["xlsx"])
    file_time = st.file_uploader("Moving time.xlsx", type=["xlsx"])
    file_steps = st.file_uploader("Steps.xlsx", type=["xlsx"])
with col2:
    file_league = st.file_uploader("league.xlsx", type=["xlsx"])
    file_move = st.file_uploader("move+.xlsx", type=["xlsx"])
    file_recharger = st.file_uploader("recharger move.xlsx", type=["xlsx"])


# ==========================================
# LOGIKA MESIN OTOMASI
# ==========================================
proses_btn = st.button("Proses Sinkronisasi Data")

if proses_btn:
    if not all([master_file, file_active, file_kcal, file_time, file_steps, file_league, file_move, file_recharger]):
        st.error("Semua 8 file wajib diunggah sebelum memproses data.")
    else:
        with st.spinner("Sinkronisasi data sedang berjalan..."):
            try:
                def to_str(x):
                    try: return str(int(float(x)))
                    except: return str(x).strip()

                # --- BACA DATA MENTAH ---
                df_active = pd.read_excel(file_active)
                df_kcal = pd.read_excel(file_kcal, header=1)
                df_time = pd.read_excel(file_time, header=1)
                df_steps = pd.read_excel(file_steps, header=1)
                df_league = pd.read_excel(file_league, header=1)
                df_move = pd.read_excel(file_move, header=1)
                df_recharger = pd.read_excel(file_recharger, header=1)

                kcal_map = {to_str(k): str(v).lower().replace(' kcal','') for k, v in zip(df_kcal['Ebib'], df_kcal['Total kcal'])}
                time_map = {to_str(k): str(v).lower().replace(' min','').replace(' minutes','') for k, v in zip(df_time['Ebib'], df_time['Total Moving Time'])}
                step_map = {to_str(k): v for k, v in zip(df_steps['Ebib'], df_steps['Total Steps'])}
                active_map = {to_str(k): v for k, v in zip(df_active['Ebib'], df_active['Active'])}
                
                league_status = {to_str(k): v for k, v in zip(df_league['EBIB'], df_league['STATUS'])}
                league_kcal = {to_str(k): v for k, v in zip(df_league['EBIB'], df_league['TOTAL KCAL'])}
                league_time = {to_str(k): v for k, v in zip(df_league['EBIB'], df_league['TOTAL MOVING TIME'])}
                league_score = {to_str(k): v for k, v in zip(df_league['EBIB'], df_league['TOTAL SCORE'])}
                league_carbon = {to_str(k): str(v).lower().replace(' gco₂e','') for k, v in zip(df_league['EBIB'], df_league['TOTAL CARBON SAVED'])}

                move_dist = {to_str(k): str(v).lower().replace(' min','') for k, v in zip(df_move['Ebib'], df_move['Distances'])}
                move_status = {to_str(k): v for k, v in zip(df_move['Ebib'], df_move['Status'])}

                rech_dist = {to_str(k): str(v).lower().replace(' poin','') for k, v in zip(df_recharger['Ebib'], df_recharger['Distances'])}
                rech_status = {to_str(k): v for k, v in zip(df_recharger['Ebib'], df_recharger['Status'])}

                wb = openpyxl.load_workbook(master_file)
                
                # --- 1. UPDATE SHEET DATA RINCIAN ---
                ws_kinerja = wb['Monit Kinerja']
                for row in range(5, ws_kinerja.max_row + 1):
                    ebib_val = ws_kinerja.cell(row=row, column=10).value
                    if ebib_val is None: continue
                    ebib_str = to_str(ebib_val)
                    if ebib_str in kcal_map:
                        v = kcal_map[ebib_str]
                        ws_kinerja.cell(row=row, column=11).value = float(v) if v.replace('.','',1).isdigit() else v
                    if ebib_str in time_map:
                        v = time_map[ebib_str]
                        ws_kinerja.cell(row=row, column=12).value = float(v) if v.replace('.','',1).isdigit() else v
                    if ebib_str in step_map: ws_kinerja.cell(row=row, column=13).value = step_map[ebib_str]
                    if ebib_str in active_map: ws_kinerja.cell(row=row, column=14).value = active_map[ebib_str]

                ws_league = wb['Monit League']
                for row in range(5, ws_league.max_row + 1):
                    ebib_val = ws_league.cell(row=row, column=8).value
                    if ebib_val is None: continue
                    ebib_str = to_str(ebib_val)
                    if ebib_str in league_status: ws_league.cell(row=row, column=10).value = league_status[ebib_str]
                    if ebib_str in league_score: ws_league.cell(row=row, column=11).value = league_score[ebib_str]
                    if ebib_str in league_kcal: ws_league.cell(row=row, column=12).value = league_kcal[ebib_str]
                    if ebib_str in league_time: ws_league.cell(row=row, column=13).value = league_time[ebib_str]
                    if ebib_str in league_carbon: ws_league.cell(row=row, column=16).value = league_carbon[ebib_str]

                ws_move = wb['Monit Move+']
                for row in range(5, ws_move.max_row + 1):
                    ebib_val = ws_move.cell(row=row, column=2).value
                    if ebib_val is None: continue
                    ebib_str = to_str(ebib_val)
                    if ebib_str in move_dist: ws_move.cell(row=row, column=6).value = move_dist[ebib_str]
                    if ebib_str in move_status: ws_move.cell(row=row, column=7).value = move_status[ebib_str]

                ws_rech = wb['Recharger Move']
                for row in range(5, ws_rech.max_row + 1):
                    ebib_val = ws_rech.cell(row=row, column=9).value
                    if ebib_val is None: continue
                    ebib_str = to_str(ebib_val)
                    if ebib_str in rech_dist: ws_rech.cell(row=row, column=10).value = rech_dist[ebib_str]
                    if ebib_str in rech_status: ws_rech.cell(row=row, column=11).value = rech_status[ebib_str]

                # --- 2. HITUNG SEMUA TABEL REKAP OTOMATIS ---
                ws_rekap = wb['All Rekap']

                # A. KINERJA
                data_k = []
                for row in ws_kinerja.iter_rows(min_row=5, values_only=True):
                    if row[5]: data_k.append({'UNIT': row[5], 'Kcal': row[10] or 0, 'Time': row[11] or 0, 'Active': row[13]})
                df_k = pd.DataFrame(data_k)
                rkp_k = df_k.groupby('UNIT').agg(
                    Total_Pegawai=('UNIT', 'count'), AVG_Kalori=('Kcal', 'mean'), AVG_Time=('Time', 'mean'),
                    Active=('Active', lambda x: (x == 'Active').sum()), Inactive=('Active', lambda x: (x == 'Inactive').sum())
                ).reset_index()
                rkp_k_dict = rkp_k.set_index('UNIT').to_dict('index')

                for row in range(6, 15):
                    unit = ws_rekap.cell(row=row, column=1).value
                    if unit and unit in rkp_k_dict:
                        d = rkp_k_dict[unit]
                        ws_rekap.cell(row=row, column=2).value = d['Total_Pegawai']
                        ws_rekap.cell(row=row, column=3).value = d['AVG_Kalori']
                        ws_rekap.cell(row=row, column=4).value = d['AVG_Time']
                        ws_rekap.cell(row=row, column=5).value = d['AVG_Kalori'] / 9000
                        ws_rekap.cell(row=row, column=6).value = d['Active']
                        ws_rekap.cell(row=row, column=7).value = d['Inactive']
                        ws_rekap.cell(row=row, column=9).value = d['Active'] / d['Total_Pegawai']
                
                tot_pegawai = rkp_k['Total_Pegawai'].sum()
                ws_rekap.cell(row=15, column=5).value = df_k['Kcal'].mean() / 9000
                ws_rekap.cell(row=15, column=9).value = rkp_k['Active'].sum() / tot_pegawai if tot_pegawai > 0 else 0

                # B. PLN LEAGUE
                data_l = []
                for row in ws_league.iter_rows(min_row=5, values_only=True):
                    if row[4]:
                        kal_pct = row[16] if isinstance(row[16], (int, float)) else 0
                        wak_pct = row[17] if isinstance(row[17], (int, float)) else 0
                        data_l.append({'UNIT': row[4], 'Kalori_pct': kal_pct, 'Waktu_pct': wak_pct})
                df_l = pd.DataFrame(data_l)
                rkp_l = df_l.groupby('UNIT').mean().reset_index()
                rkp_l_dict = rkp_l.set_index('UNIT').to_dict('index')

                for row in range(23, 32):
                    unit = ws_rekap.cell(row=row, column=1).value
                    if unit and unit in rkp_l_dict:
                        ws_rekap.cell(row=row, column=2).value = rkp_l_dict[unit]['Kalori_pct']
                        ws_rekap.cell(row=row, column=3).value = rkp_l_dict[unit]['Waktu_pct']
                
                ws_rekap.cell(row=32, column=2).value = df_l['Kalori_pct'].mean()
                ws_rekap.cell(row=32, column=3).value = df_l['Waktu_pct'].mean()

                # C. PLN RECHARGER
                data_r = []
                for row in ws_rech.iter_rows(min_row=5, values_only=True):
                    if row[4]:
                        try: dist = float(str(row[9]).lower().replace(' poin','').strip())
                        except: dist = 0
                        data_r.append({'UNIT': row[4], 'Status': row[10], 'Distances': dist})
                df_r = pd.DataFrame(data_r)
                rkp_r = df_r.groupby('UNIT').agg(
                    Finish=('Status', lambda x: (x == 'Finish').sum()),
                    Ongoing=('Status', lambda x: (x == 'Ongoing').sum()), Distances=('Distances', 'sum')
                ).reset_index()
                rkp_r['Grand Total'] = rkp_r['Finish'] + rkp_r['Ongoing']
                rkp_r_dict = rkp_r.set_index('UNIT').to_dict('index')

                for row in range(37, 46):
                    unit = ws_rekap.cell(row=row, column=1).value
                    if unit and unit in rkp_r_dict:
                        d = rkp_r_dict[unit]
                        ws_rekap.cell(row=row, column=2).value = d['Finish']
                        ws_rekap.cell(row=row, column=3).value = d['Ongoing']
                        ws_rekap.cell(row=row, column=4).value = d['Distances']
                        ws_rekap.cell(row=row, column=5).value = d['Grand Total']
                        ws_rekap.cell(row=row, column=6).value = d['Finish'] / d['Grand Total'] if d['Grand Total'] > 0 else 0
                
                tot_f = rkp_r['Finish'].sum()
                tot_o = rkp_r['Ongoing'].sum()
                gt_r = tot_f + tot_o
                ws_rekap.cell(row=46, column=2).value = tot_f
                ws_rekap.cell(row=46, column=3).value = tot_o
                ws_rekap.cell(row=46, column=4).value = rkp_r['Distances'].sum()
                ws_rekap.cell(row=46, column=5).value = gt_r
                ws_rekap.cell(row=46, column=6).value = tot_f / gt_r if gt_r > 0 else 0

                # D. REKAP MOVE+ SUMMARY
                data_m = []
                for row in ws_move.iter_rows(min_row=5, values_only=True):
                    if row[11] and row[11] != 'UNIT': data_m.append({'UNIT': row[11], 'Status': str(row[6]).strip()})
                df_m = pd.DataFrame(data_m)
                rkp_m = df_m.groupby('UNIT').agg(
                    Min=('Status', lambda x: (x == '-').sum()), Finish=('Status', lambda x: (x == 'Finish').sum()),
                    Ongoing=('Status', lambda x: (x == 'Ongoing').sum())
                ).reset_index()
                rkp_m['Grand Total'] = rkp_m['Min'] + rkp_m['Finish'] + rkp_m['Ongoing']
                rkp_m_dict = rkp_m.set_index('UNIT').to_dict('index')

                ws_move_rekap = wb['Rekap Move+']
                for row in range(5, 14):
                    unit = ws_move_rekap.cell(row=row, column=1).value
                    if unit and unit in rkp_m_dict:
                        d = rkp_m_dict[unit]
                        ws_move_rekap.cell(row=row, column=2).value = d['Min']
                        ws_move_rekap.cell(row=row, column=3).value = d['Finish']
                        ws_move_rekap.cell(row=row, column=4).value = d['Ongoing']
                        ws_move_rekap.cell(row=row, column=5).value = d['Grand Total']
                        ws_move_rekap.cell(row=row, column=6).value = (d['Finish'] + d['Ongoing']) / d['Grand Total'] if d['Grand Total'] > 0 else 0

                tot_min, tot_fin, tot_ong, gt_m = rkp_m['Min'].sum(), rkp_m['Finish'].sum(), rkp_m['Ongoing'].sum(), rkp_m['Grand Total'].sum()
                ws_move_rekap.cell(row=14, column=2).value = tot_min
                ws_move_rekap.cell(row=14, column=3).value = tot_fin
                ws_move_rekap.cell(row=14, column=4).value = tot_ong
                ws_move_rekap.cell(row=14, column=5).value = gt_m
                ws_move_rekap.cell(row=14, column=6).value = (tot_fin + tot_ong) / gt_m if gt_m > 0 else 0

                # =========================================================
                # E. KLASEMEN TOP 3 MOVE+
                # =========================================================
                data_m_full = []
                for row in ws_move.iter_rows(min_row=5, values_only=True):
                    if row[0]:
                        dist_val = row[5] 
                        try: dist_num = float(str(dist_val).lower().replace(' min','').replace(',','.').strip())
                        except: dist_num = 0
                        
                        data_m_full.append({
                            'Name': row[2], 'NAMA': row[3], 'Category': row[4], 
                            'Distances': row[5], 'Distances_num': dist_num, 'Status': row[6],
                            'Join Event': row[7], 'USIA': row[8], 'Kategori Usia': row[9],
                            'GENDER': row[10], 'UNIT': row[11], 'SUB UNIT': row[12]
                        })
                df_move_full = pd.DataFrame(data_m_full)

                klasemen_rules = [
                    ('80 Km Run', 'Male', 'Usia <45 Tahun'),
                    ('80 Km Run', 'Male', 'usia >=45 Tahun'),
                    ('80 Km Run', 'Female', 'Usia <45 Tahun'),
                    ('60 Km Walk', 'Male', 'Usia <45 Tahun'),
                    ('60 Km Walk', 'Male', 'usia >=45 Tahun'),
                    ('60 Km Walk', 'Female', 'Usia <45 Tahun'),
                    ('60 Km Walk', 'Female', 'usia >=45 Tahun'),
                    ('500 Km Ride', 'Male', None),
                    ('2700 Minutes', 'Male', 'Usia <45 Tahun'),
                    ('2700 Minutes', 'Male', 'usia >=45 Tahun'),
                    ('2700 Minutes', 'Female', 'Usia <45 Tahun'),
                    ('2700 Minutes', 'Female', 'usia >=45 Tahun')
                ]

                header_rows = []
                for r in range(15, ws_move_rekap.max_row):
                    if ws_move_rekap.cell(row=r, column=1).value == 'Name':
                        header_rows.append(r)

                for idx, rule in enumerate(klasemen_rules):
                    if idx < len(header_rows):
                        r = header_rows[idx]
                        cat, gen, kat_usia = rule
                        
                        if kat_usia is None: 
                            subset = df_move_full[
                                (df_move_full['Category'].astype(str).str.strip().str.lower() == cat.lower()) &
                                (df_move_full['GENDER'].astype(str).str.strip().str.lower() == gen.lower())
                            ].copy()
                        else:
                            subset = df_move_full[
                                (df_move_full['Category'].astype(str).str.strip().str.lower() == cat.lower()) &
                                (df_move_full['GENDER'].astype(str).str.strip().str.lower() == gen.lower()) &
                                (df_move_full['Kategori Usia'].astype(str).str.strip().str.lower() == kat_usia.lower())
                            ].copy()
                        
                        subset = subset.sort_values(by='Distances_num', ascending=False).head(3)
                        
                        for i in range(3):
                            target_row = r + 1 + i
                            if i < len(subset):
                                record = subset.iloc[i]
                                ws_move_rekap.cell(row=target_row, column=1).value = record['Name']
                                ws_move_rekap.cell(row=target_row, column=2).value = record['NAMA']
                                ws_move_rekap.cell(row=target_row, column=3).value = record['Category']
                                ws_move_rekap.cell(row=target_row, column=4).value = record['Distances']
                                ws_move_rekap.cell(row=target_row, column=5).value = record['Status']
                                ws_move_rekap.cell(row=target_row, column=6).value = record['Join Event']
                                ws_move_rekap.cell(row=target_row, column=7).value = record['USIA']
                                ws_move_rekap.cell(row=target_row, column=8).value = record['Kategori Usia']
                                ws_move_rekap.cell(row=target_row, column=9).value = record['GENDER']
                                ws_move_rekap.cell(row=target_row, column=10).value = record['UNIT']
                                ws_move_rekap.cell(row=target_row, column=11).value = record['SUB UNIT']
                            else:
                                for col_idx in [1, 2, 4, 5, 6, 7, 10, 11]:
                                    ws_move_rekap.cell(row=target_row, column=col_idx).value = None

                # --- 3. SAVE & DOWNLOAD ---
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("✅ Sinkronisasi Selesai!")
                st.download_button(
                    label="Download Excel",
                    data=output,
                    file_name="Monitoring_PLN_Wellness_UPDATED_Final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            except Exception as e:
                st.error(f"Terjadi kendala: {e}")

# ==========================================
# FOOTER SIGNATURE
# ==========================================
st.markdown("""
    <div class="footer-text">
        developed by <span>msalmannal</span>
    </div>
""", unsafe_allow_html=True)